import os
import json
import segno
from PIL import Image, ImageTk
from tkinter import Tk, Label, Button, filedialog, Canvas, Scale, HORIZONTAL, StringVar, Entry, OptionMenu
from tkinter.messagebox import showinfo, showerror
from openpyxl import load_workbook

THEME_DIR = "themes"
os.makedirs(THEME_DIR, exist_ok=True)

# 預設主題
default_theme = {
    "fill": "#000000",
    "background": None,  # None 代表透明背景
    "style": "round",    # 支援 round, square, circle-modules
    "scale": 10,
    "border": 2
}

# 儲存主題到 JSON
def save_theme(theme_name, theme):
    with open(f"{THEME_DIR}/{theme_name}.json", "w") as f:
        json.dump(theme, f)

# 載入所有主題
def load_all_themes():
    themes = {}
    for file in os.listdir(THEME_DIR):
        if file.endswith(".json"):
            with open(f"{THEME_DIR}/{file}", "r") as f:
                name = file.replace(".json", "")
                themes[name] = json.load(f)
    return themes

# 產生 QR 圖像（返回 PIL.Image）
def generate_qr_image(data, theme, logo_path=None, logo_size_ratio=0.2, logo_offset=(0, 0)):
    qr = segno.make(data, error='h')
    qr_img_path = "_temp.png"
    qr.save(qr_img_path,
            scale=theme.get("scale", 10),
            border=theme.get("border", 2),
            color=theme.get("fill", "#000"),
            background=theme.get("background", None),
            kind=theme.get("style", "round"))
    base = Image.open(qr_img_path).convert("RGBA")
    os.remove(qr_img_path)

    if logo_path and os.path.exists(logo_path):
        logo = Image.open(logo_path).convert("RGBA")
        lw = int(base.width * logo_size_ratio)
        logo = logo.resize((lw, lw))
        x = (base.width - lw) // 2 + logo_offset[0]
        y = (base.height - lw) // 2 + logo_offset[1]
        base.paste(logo, (x, y), logo)

    return base

# 預覽區更新函數
def update_preview():
    data = "https://example.com"
    logo_ratio = logo_size_slider.get() / 100
    offset = (int(offset_x.get()), int(offset_y.get()))
    try:
        img = generate_qr_image(data, current_theme, logo_path_var.get(), logo_ratio, offset)
        preview_img = ImageTk.PhotoImage(img.resize((200, 200)))
        canvas.delete("all")
        canvas.create_image(100, 100, image=preview_img)
        canvas.image = preview_img
    except Exception as e:
        print("預覽錯誤:", e)

# 批次從 Excel 產生 QR Code
def batch_generate_qr():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return
    output_folder = filedialog.askdirectory(title="Choose Output Folder")
    if not output_folder:
        return

    wb = load_workbook(file_path)
    sheet = wb.active
    logo_ratio = logo_size_slider.get() / 100
    offset = (int(offset_x.get()), int(offset_y.get()))
    count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        filename, url = row[:2]
        if not filename or not url:
            continue
        try:
            img = generate_qr_image(url, current_theme, logo_path_var.get(), logo_ratio, offset)
            out_path = os.path.join(output_folder, f"{filename}.png")
            img.save(out_path)
            count += 1
        except Exception as e:
            print(f"Error on {filename}: {e}")

    showinfo("Done", f"{count} QR codes generated!")

# 選擇 Logo
def select_logo():
    path = filedialog.askopenfilename(filetypes=[("Image", "*.png;*.jpg;*.jpeg")])
    if path:
        logo_path_var.set(path)
        update_preview()

# 儲存主題
def save_theme_action():
    name = theme_name_var.get()
    if not name:
        showerror("Error", "Please enter a theme name.")
        return
    save_theme(name, current_theme)
    showinfo("Saved", f"Theme '{name}' saved!")
    refresh_theme_menu()

# 切換主題
def load_theme(name):
    global current_theme
    current_theme = all_themes[name]
    update_preview()

# 重新整理主題清單
def refresh_theme_menu():
    global all_themes
    all_themes = load_all_themes()
    menu = theme_dropdown["menu"]
    menu.delete(0, "end")
    for name in all_themes:
        menu.add_command(label=name, command=lambda value=name: load_theme(value))

# GUI 界面
root = Tk()
root.title("QR Code Generator - Advanced")
root.geometry("600x600")

canvas = Canvas(root, width=200, height=200)
canvas.pack(pady=10)

logo_path_var = StringVar()
theme_name_var = StringVar()
offset_x = StringVar(value="0")
offset_y = StringVar(value="0")

Button(root, text="Choose Logo", command=select_logo).pack()
Scale(root, from_=10, to=50, orient=HORIZONTAL, label="Logo Size (%)", command=lambda e: update_preview()).pack()
logo_size_slider = root.children[list(root.children)[-1]]

Label(root, text="Offset X:").pack()
Entry(root, textvariable=offset_x).pack()
Label(root, text="Offset Y:").pack()
Entry(root, textvariable=offset_y).pack()

Label(root, text="Theme name:").pack()
Entry(root, textvariable=theme_name_var).pack()
Button(root, text="Save Theme", command=save_theme_action).pack()

all_themes = load_all_themes()
theme_list = list(all_themes.keys()) or ["default"]
current_theme = all_themes.get(theme_list[0], default_theme)

theme_dropdown_var = StringVar(value=theme_list[0])
theme_dropdown = OptionMenu(root, theme_dropdown_var, *theme_list, command=load_theme)
theme_dropdown.pack()

Button(root, text="Generate QR from Excel", command=batch_generate_qr).pack(pady=10)

update_preview()
root.mainloop()
