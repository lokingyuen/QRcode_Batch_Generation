import os
import qrcode
from tkinter import Tk, filedialog, Button, Label
from openpyxl import load_workbook
from PIL import Image
import tkinter.messagebox as messagebox
import threading

# Function to generate QR codes
def generate_qr_code(excel_file, output_folder):
    try:
        # Read the Excel file
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Iterate through each row in Excel, assuming the first column is the image name, and the second column is the link
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            file_name = row[0].value  # First column is the file name
            link = row[1].value  # Second column is the link

            if file_name and link:
                # Generate QR code
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
                qr.add_data(link)
                qr.make(fit=True)
                img = qr.make_image(fill='black', back_color='white')

                # Ensure the output folder exists
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                # Save the QR code as a file
                img.save(os.path.join(output_folder, f"{file_name}.png"))

        # Show a message when the generation is complete
        messagebox.showinfo("Completed", "QR code generation successful!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred during the generation process: {e}")

# Open the file selection dialog
def open_file():
    # Let the user choose an Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        excel_label.config(text=f"Selected Excel File: {file_path}")  # Show the selected Excel file
        # Let the user choose the output folder
        output_folder = filedialog.askdirectory(title="Select Output Folder")
        if output_folder:
            folder_label.config(text=f"Selected Output Folder: {output_folder}")  # Show the selected output folder
            start_button.config(state="normal", command=lambda: start_generation(file_path, output_folder))

# Start generating QR codes
def start_generation(excel_file, output_folder):
    # Create a new thread to run the generation process so that it doesn't block the main interface
    generation_thread = threading.Thread(target=generate_qr_code, args=(excel_file, output_folder))
    generation_thread.start()

# Create the GUI interface
def create_gui():
    global excel_label, folder_label, start_button
    
    root = Tk()
    root.title("QR Code Generator Powered by King")  # Title of the window
    root.geometry("500x350")
    
    label = Label(root, text="Select an Excel file to generate QR codes", font=("Arial", 12))
    label.pack(pady=20)

    open_button = Button(root, text="Open Excel File", command=open_file, font=("Arial", 12))
    open_button.pack(pady=20)

    excel_label = Label(root, text="Selected Excel File: Not Selected", font=("Arial", 10))
    excel_label.pack(pady=10)
    
    folder_label = Label(root, text="Selected Output Folder: Not Selected", font=("Arial", 10))
    folder_label.pack(pady=10)
    
    start_button = Button(root, text="Batch Generation", state="disabled", font=("Arial", 12))
    start_button.pack(pady=20)

    # Display the footer text
    bottom_label = Label(root, text="Friendships never go out of style, But betrayal will", font=("Arial", 10), fg="lightgrey")
    bottom_label.pack(side="bottom", pady=10)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
