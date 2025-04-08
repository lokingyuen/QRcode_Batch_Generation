import os
import qrcode
from tkinter import Tk, filedialog, Button, Label
from openpyxl import load_workbook
from PIL import Image
import tkinter.messagebox as messagebox

# 定義生成QR代碼的函數
def generate_qr_code(excel_file, output_folder):
    try:
        # 讀取 Excel 文件
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # 遍歷 Excel 中的每一行，假設第一列是圖片名稱，第二列是鏈接
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            file_name = row[0].value  # 第一列是文件名稱
            link = row[1].value  # 第二列是鏈接

            if file_name and link:
                # 生成 QR 代碼
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
                qr.add_data(link)
                qr.make(fit=True)
                img = qr.make_image(fill='black', back_color='white')

                # 確保輸出文件夾存在
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                # 保存為文件
                img.save(os.path.join(output_folder, f"{file_name}.png"))
        
        messagebox.showinfo("完成", "QR 代碼生成成功！")
    except Exception as e:
        messagebox.showerror("錯誤", f"生成過程中出現錯誤: {e}")

# 打開文件選擇對話框
def open_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if file_path:
        output_folder = filedialog.askdirectory(title="選擇保存文件夾")
        if output_folder:
            generate_qr_code(file_path, output_folder)

# 創建GUI界面
def create_gui():
    root = Tk()
    root.title("QR Code Generator")
    root.geometry("400x200")
    
    label = Label(root, text="選擇一個 Excel 文件來生成 QR 代碼", font=("Arial", 12))
    label.pack(pady=20)

    open_button = Button(root, text="打開 Excel 文件", command=open_file, font=("Arial", 12))
    open_button.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    create_gui()
